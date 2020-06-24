import math
import openpyxl
import os
from shutil import copyfile
from xmindparser import xmind_to_dict


def get_cases_list_from_xmind_file(xmind_file):
    # 获取需求数据
    story_data = xmind_to_dict(xmind_file)[0]['topic']['topics']
    # 创建包含文件中所有用例的列表
    case_list = []
    for story in story_data:
        # 获取模块数据
        module_data = story['topics']
        for module in module_data:
            # 获取目录数据
            index_data = module['topics']
            for index in index_data:
                # 获取用例数据
                case_data = index['topics']
                for case in case_data:
                    # 创建单个用例的数据列表，并添加需求名称，模块名称，目录名称和用例名称
                    lst = [story['title'], module['title'], index['title'], case['title']]
                    # 获取用例内容
                    case_content_data = case['topics']
                    # 添加前置条件和级别到用例列表中
                    lst.append(case_content_data[0]['title'])
                    lst.append(case_content_data[1]['title'])
                    for i in range(0, len(case_content_data)-2):
                        # 获取用例步骤并添加到用例列表中
                        step_data = case_content_data[i+2]
                        lst.append(step_data['title'])
                        # 获取用例执行结果并添加到用例列表中
                        result_data = step_data['topics']
                        for result in result_data:
                            lst.append(result['title'])
                    # 将该条用例添加到所有用例列表中
                    case_list.append(lst)
    return case_list


def covert_to_excel(templatename, lst, filename, count):
    # 超过500条用例需要分文件处理
    file_count = int(math.ceil(len(lst)/count))
    print("---共获取到"+str(len(lst))+"条用例，将会写入到"+str(file_count)+"个文件中")
    for c in range(0, file_count):
        copyfile(templatename, filename+"_测试用例导入_"+str(c+1)+".xlsx")
        wb = openpyxl.load_workbook(filename+"_测试用例导入_"+str(c+1)+".xlsx")
        xl_sheet = wb.get_sheet_by_name("sheet1")
        for i in range(c*count, c*count+count):
            if i == len(lst):
                break
            print("------正在写入第"+str(i+1)+"条用例...")
            case_to_excel = lst[i]
            for j in range(0, len(case_to_excel)):
                if j == 0:
                    xl_sheet.cell(row=i % count+2, column=9).value = str_split(case_to_excel[j])[1]
                    xl_sheet.cell(row=i % count+2, column=10).value = str_split(case_to_excel[j])[2]
                elif j == 1:
                    xl_sheet.cell(row=i % count+2, column=8).value = str_split(case_to_excel[j])[1]
                elif j == 2:
                    xl_sheet.cell(row=i % count+2, column=13).value = str_split(case_to_excel[j])[1]
                elif j == 3:
                    xl_sheet.cell(row=i % count+2, column=1).value = str_split(case_to_excel[j])[1]
                elif j == 4:
                    xl_sheet.cell(row=i % count+2, column=6).value = str_split(case_to_excel[j])[1]
                elif j == 5:
                    xl_sheet.cell(row=i % count+2, column=12).value = str_split(case_to_excel[j])[1]
                else:
                    xl_sheet.cell(row=i % count + 2, column=j + 8).value = str_split(case_to_excel[j])[1]

        wb.save(filename+"_测试用例导入_"+str(c+1)+".xlsx")


def str_split(str_data):
    return str_data.split("##", 2)


if __name__ == '__main__':
    # 获取当前文档路径
    path = os.getcwd()
    # 设置华为云用例模板Excel名称
    excel_template_name = "importTemplate-20200615201950.xlsx"
    # 获取当前文档下的xmind文件名称
    files = os.listdir(path)
    for file in files:
        if os.path.splitext(file)[1] == ".xmind":
            xmind_file_name = file
            print("正在转换"+file+"中的用例：")
            cases_list = get_cases_list_from_xmind_file(os.path.join(path, xmind_file_name))
            # 从xmind中获取用例文件名称
            cases_file_name = os.path.splitext(file)[0]
            # 转化到Excel中，需要填写Excel模板名字，获取到的用例数据，获取到的用例名字，以及每个文件允许最大的用例个数，华为云默认500
            covert_to_excel(os.path.join(path, excel_template_name), cases_list, cases_file_name, 500)
